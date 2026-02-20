"""DORA compliance API — register, coverage, gaps, export."""

import io
import uuid
from datetime import date, datetime, timezone
from typing import Any

from fastapi import APIRouter
from fastapi.responses import StreamingResponse

from app.schemas.compliance import (
    DORAcoverage,
    DORAGap,
    DORARegisterCreate,
    DORARegisterEntry,
)

router = APIRouter(prefix="/compliance", tags=["compliance"])

# In-memory store (replaced by DB in production)
_register: dict[str, DORARegisterEntry] = {}

# Seed demo data
_DEMO_ENTRIES = [
    DORARegisterEntry(
        id="dora-001", vendor_id="v-001", vendor_name="CloudCorp SAS",
        service_type="IaaS Cloud", critical=True, tier=1,
        contract_start=date(2024, 1, 1), contract_end=date(2026, 12, 31),
        last_audit=date(2025, 11, 15), score=820, grade="A",
        compliance_status="conforme",
    ),
    DORARegisterEntry(
        id="dora-002", vendor_id="v-002", vendor_name="SecurePay Ltd",
        service_type="Payment Processing", critical=True, tier=1,
        contract_start=date(2023, 6, 1), contract_end=date(2026, 5, 31),
        last_audit=date(2025, 9, 1), score=650, grade="B",
        compliance_status="partiel",
    ),
    DORARegisterEntry(
        id="dora-003", vendor_id="v-003", vendor_name="DataBack GmbH",
        service_type="Backup & DR", critical=False, tier=2,
        contract_start=date(2024, 3, 1), contract_end=date(2027, 2, 28),
        last_audit=date(2025, 6, 20), score=540, grade="C",
        compliance_status="partiel",
    ),
    DORARegisterEntry(
        id="dora-004", vendor_id="v-004", vendor_name="MailGuard Inc",
        service_type="Email Security", critical=False, tier=3,
        contract_start=date(2025, 1, 1), contract_end=date(2027, 12, 31),
        last_audit=None, score=0, grade="F",
        compliance_status="non-conforme",
    ),
]
for e in _DEMO_ENTRIES:
    _register[e.id] = e


@router.get("/dora/register", response_model=list[DORARegisterEntry])
async def get_dora_register() -> list[DORARegisterEntry]:
    """Return the full DORA ICT provider register."""
    return list(_register.values())


@router.post("/dora/register", response_model=DORARegisterEntry)
async def add_dora_entry(body: DORARegisterCreate) -> DORARegisterEntry:
    """Add or update a DORA register entry."""
    entry_id = str(uuid.uuid4())
    entry = DORARegisterEntry(
        id=entry_id,
        vendor_id=body.vendor_id,
        vendor_name=body.vendor_id,  # In production, resolve from vendor DB
        service_type=body.service_type,
        critical=body.critical,
        tier=body.tier,
        contract_start=body.contract_start,
        contract_end=body.contract_end,
        notes=body.notes,
    )
    _register[entry_id] = entry
    return entry


@router.get("/dora/coverage", response_model=DORAcoverage)
async def get_dora_coverage() -> DORAcoverage:
    """Return DORA register coverage statistics."""
    entries = list(_register.values())
    total = len(entries)
    critical = sum(1 for e in entries if e.critical)
    tier1 = sum(1 for e in entries if e.tier == 1)
    tier2 = sum(1 for e in entries if e.tier == 2)
    tier3 = sum(1 for e in entries if e.tier == 3)
    evaluated = sum(1 for e in entries if e.last_audit is not None)
    # Coverage = entries with score > 0 / total
    scored = sum(1 for e in entries if e.score > 0)
    coverage = (scored / total * 100) if total > 0 else 0

    return DORAcoverage(
        total_vendors=total,
        registered=total,
        coverage_pct=round(coverage, 1),
        critical_count=critical,
        tier1_count=tier1,
        tier2_count=tier2,
        tier3_count=tier3,
        evaluated_this_quarter=evaluated,
        last_updated=datetime.now(timezone.utc),
    )


@router.get("/dora/gaps", response_model=list[DORAGap])
async def get_dora_gaps() -> list[DORAGap]:
    """Identify gaps in DORA compliance."""
    gaps: list[DORAGap] = []
    for entry in _register.values():
        if entry.last_audit is None:
            gaps.append(DORAGap(
                id=f"gap-audit-{entry.id}",
                vendor_id=entry.vendor_id,
                vendor_name=entry.vendor_name,
                gap_type="missing_audit",
                description=f"No audit recorded for {entry.vendor_name}",
                severity="high",
                recommendation="Schedule an initial security audit.",
            ))
        if entry.compliance_status == "non-conforme":
            gaps.append(DORAGap(
                id=f"gap-compliance-{entry.id}",
                vendor_id=entry.vendor_id,
                vendor_name=entry.vendor_name,
                gap_type="non_compliant",
                description=f"{entry.vendor_name} is non-compliant",
                severity="critical",
                recommendation="Initiate remediation plan and escalation.",
            ))
        if entry.critical and entry.score < 600:
            gaps.append(DORAGap(
                id=f"gap-score-{entry.id}",
                vendor_id=entry.vendor_id,
                vendor_name=entry.vendor_name,
                gap_type="low_score_critical",
                description=f"Critical provider {entry.vendor_name} has low score ({entry.score})",
                severity="high",
                recommendation="Engage vendor for immediate remediation.",
            ))
    return gaps


@router.get("/dora/export")
async def export_dora_register() -> StreamingResponse:
    """Export the DORA register as an Excel file."""
    try:
        from openpyxl import Workbook
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "DORA Register"

    headers = [
        "ID", "Vendor", "Service Type", "Critical", "Tier",
        "Contract Start", "Contract End", "Last Audit",
        "Score", "Grade", "Compliance Status", "Notes",
    ]
    ws.append(headers)

    for entry in _register.values():
        ws.append([
            entry.id,
            entry.vendor_name,
            entry.service_type,
            "Yes" if entry.critical else "No",
            entry.tier,
            str(entry.contract_start) if entry.contract_start else "",
            str(entry.contract_end) if entry.contract_end else "",
            str(entry.last_audit) if entry.last_audit else "",
            entry.score,
            entry.grade,
            entry.compliance_status,
            entry.notes,
        ])

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dora_register.xlsx"},
    )
